"""
Hisse Senedi Temel Veri Çekici — GitHub Actions Uyumlu (Tam Sürüm)
==================================================================
yfinance üzerinden ulaşılabilen tüm temel finansal verileri çeker ve
sonuçları kategorilere ayrılmış, biçimlendirilmiş Excel dosyasına kaydeder.
Her sütun başlığında açıklayıcı not (hover ile görünür) bulunur.

Hisse listesi aynı dizindeki 'hisseisimleri.txt' dosyasından alınır.
Çıktı dosyası adı otomatik olarak gün_ay_yıl içerir (ör: hisse_temel_veriler_21_03_2025.xlsx)

Kurulum:
    pip install yfinance pandas openpyxl requests

Kullanım:
    python temel_verileri_cek.py
"""

import sys
import time
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta


# ─── AYARLAR ──────────────────────────────────────────────────────────────────
HISSE_LISTESI_DOSYASI = "hisseisimleri.txt"   # aynı dizinde olmalı
ISTEK_ARASI_BEKLEME = 0.6                     # saniye — API'yi yormamak için
# ──────────────────────────────────────────────────────────────────────────────


# ─── SÜTUN TANIMI ─────────────────────────────────────────────────────────────
# (başlık, info_anahtarı, ön_işlem, number_format, excel_notu)
# ön_işlem : None | "yuzde" | "tarih" | "ozel"

SUTUNLAR = [

    # ── 1. KİMLİK & ŞİRKET BİLGİSİ ──────────────────────────────────────────
    (
        "Hisse Kodu", None, "ozel", None,
        "Hissenin BIST'teki işlem kodu.\n"
        "Öneri: Filtreleme ve DÜŞEYARA işlemleri için bu sütunu anahtar olarak kullanın."
    ),
    (
        "Şirket İsmi", "longName", None, None,
        "Şirketin tam ticari unvanı.\n"
        "Öneri: Kısaltılmış ad için 'shortName' alanı kullanılabilir; ikisi arasında fark varsa dikkat edin."
    ),
    (
        "Sektör", "sector", None, None,
        "Yahoo Finance'in şirkete atadığı ana sektör (örn. Financial Services, Industrials).\n"
        "Öneri: Sektör bazında gruplama yaparak ortalama F/K ve ROE karşılaştırması yapın."
    ),
    (
        "Sektör Altı", "industry", None, None,
        "Sektörün alt kolu; şirketin faaliyet alanını daha dar tanımlar.\n"
        "Öneri: Benzer iş modeline sahip rakipleri bulmak için bu alanı kullanın."
    ),
    (
        "Ülke", "country", None, None,
        "Şirketin kayıtlı olduğu ülke.\n"
        "Öneri: BIST hisseleri için 'Turkey' beklenir; farklı bir ülke çıkıyorsa veri doğruluğunu kontrol edin."
    ),
    (
        "Çalışan Sayısı", "fullTimeEmployees", None, '#,##0',
        "Tam zamanlı çalışan sayısı.\n"
        "Öneri: Gelir / Çalışan oranı hesaplayarak operasyonel verimliliği karşılaştırın.\n"
        "Not: Bazı şirketler bu veriyi Yahoo'ya bildirmeyebilir."
    ),
    (
        "Borsa", "exchange", None, None,
        "Hissenin işlem gördüğü borsa kodu (örn. IST = Borsa İstanbul).\n"
        "Öneri: Çok borsada işlem gören şirketleri tespit etmek için kullanın."
    ),

    # ── 2. FİYAT & PİYASA ────────────────────────────────────────────────────
    (
        "Güncel Fiyat", "currentPrice", None, '#,##0.00',
        "En son işlem fiyatı (TL).\n"
        "Öneri: 50G ve 200G ortalamalarla karşılaştırarak trendin üstünde mi altında mı işlem gördüğünü belirleyin.\n"
        "Not: Borsa kapalıyken bir önceki kapanış fiyatı gösterilir."
    ),
    (
        "Piyasa Değeri", "marketCap", None, '#,##0',
        "Dolaşımdaki tüm hisselerin piyasa fiyatıyla çarpımı (TL).\n"
        "Öneri: Büyük sermayeli (large-cap) ve küçük sermayeli (small-cap) hisseleri sınıflandırmak için kullanın.\n"
        "Kural: >10 milyar TL büyük; 1–10 milyar orta; <1 milyar küçük ölçekli."
    ),
    (
        "Firma Değeri", "enterpriseValue", None, '#,##0',
        "Piyasa Değeri + Net Borç; şirketi satın almanın toplam maliyeti (TL).\n"
        "Öneri: FD/FAVÖK ve FD/Gelir oranlarının paydası olarak kullanarak sermaye yapısından bağımsız değerleme yapın.\n"
        "Not: Yüksek borçlu şirketlerde Firma Değeri Piyasa Değerinin çok üzerinde olabilir."
    ),
    (
        "Dolaşımdaki Hisse", "floatShares", None, '#,##0',
        "Halka açık, serbest dolaşımdaki hisse adedi.\n"
        "Öneri: Float düşükse (toplam hissenin %20'sinden azı) fiyat manipülasyonuna karşı hassasiyet artar; dikkatli olun."
    ),
    (
        "Toplam Hisse Sayısı", "sharesOutstanding", None, '#,##0',
        "Tedavüldeki toplam hisse adedi (hazine hisseleri dahil).\n"
        "Öneri: Float / Toplam Hisse oranı düşükse yüksek kurumsal veya içeriden sahiplik var demektir."
    ),
    (
        "52H Yüksek", "fiftyTwoWeekHigh", None, '#,##0.00',
        "Son 52 haftanın en yüksek işlem fiyatı.\n"
        "Öneri: Güncel fiyatın 52H Yüksek'e oranı (%100'e yakınsa) momentum stratejisi için olumlu sinyal olabilir."
    ),
    (
        "52H Düşük", "fiftyTwoWeekLow", None, '#,##0.00',
        "Son 52 haftanın en düşük işlem fiyatı.\n"
        "Öneri: Güncel fiyat 52H Düşük'e yakınsa değer yatırımcıları için incelemeye değer; ancak düşüşün sebebini araştırın."
    ),
    (
        "50G Ortalama", "fiftyDayAverage", None, '#,##0.00',
        "Son 50 günün kapanış fiyatı ortalaması; kısa-orta vadeli trend göstergesi.\n"
        "Öneri: Güncel fiyat > 50G Ortalama ise kısa vadeli yükseliş trendi; altındaysa baskı altında demektir."
    ),
    (
        "200G Ortalama", "twoHundredDayAverage", None, '#,##0.00',
        "Son 200 günün kapanış fiyatı ortalaması; uzun vadeli trend göstergesi.\n"
        "Öneri: 50G Ortalama > 200G Ortalama geçişi (Altın Çarpı) güçlü yükseliş sinyali;\n"
        "50G < 200G geçişi (Ölüm Çarpı) ise düşüş sinyali olarak yorumlanır."
    ),
    (
        "Ort. Hacim (3A)", "averageVolume", None, '#,##0',
        "Son 3 aylık günlük ortalama işlem hacmi (lot).\n"
        "Öneri: Günlük hacim 3A ortalamasının 2 katına çıkıyorsa önemli bir hareket veya haber olabilir."
    ),
    (
        "Ort. Hacim (10G)", "averageDailyVolume10Day", None, '#,##0',
        "Son 10 günlük günlük ortalama işlem hacmi (lot).\n"
        "Öneri: 3A ortalamasıyla karşılaştırın; 10G ortalaması belirgin şekilde yüksekse kısa vadeli ilgi artışı var demektir."
    ),
    (
        "Beta", "beta", None, '#,##0.00',
        "Hissenin endekse göre oynaklık katsayısı.\n"
        "Yorum: Beta = 1 → endeksle aynı hareket; >1 → daha oynak; <1 → daha sakin; <0 → ters hareket.\n"
        "Öneri: Düşük riskli portföy için Beta < 1 hisseleri tercih edin; yüksek getiri peşindeyseniz Beta > 1 bakın."
    ),

    # ── 3. DEĞERLEME ORANLARI ─────────────────────────────────────────────────
    (
        "F/K (Trailing)", "trailingPE", None, '#,##0.00',
        "Fiyat / Kazanç oranı — son 12 aylık gerçekleşen kazanca göre.\n"
        "Öneri: Sektör ortalamasının altındaki F/K ucuz olabilir; ancak düşük büyüme beklentisinin yansıması da olabilir.\n"
        "Kural: Negatif F/K → şirket zarar ediyor demektir, bu durumda Forward F/K'ya bakın."
    ),
    (
        "F/K (Forward)", "forwardPE", None, '#,##0.00',
        "Fiyat / Kazanç oranı — önümüzdeki 12 aylık analist kazanç tahminlerine göre.\n"
        "Öneri: Forward F/K < Trailing F/K ise kazançların büyümesi bekleniyor; bu olumlu bir sinyal.\n"
        "Not: Analist tahminlerine dayandığı için gerçekleşmeyebilir."
    ),
    (
        "PEG Oranı", "pegRatio", None, '#,##0.00',
        "F/K oranını büyüme hızına böler: PEG = F/K ÷ Kazanç Büyüme %.\n"
        "Yorum: PEG < 1 → büyümeye göre ucuz; PEG > 2 → büyümeye göre pahalı.\n"
        "Öneri: Yüksek büyüme hisselerini değerlerken F/K yerine PEG kullanmak daha sağlıklıdır."
    ),
    (
        "F/DD (P/B)", "priceToBook", None, '#,##0.00',
        "Fiyat / Defter Değeri — piyasa fiyatının özkaynak defter değerine oranı.\n"
        "Yorum: F/DD < 1 → hisse defter değerinin altında işlem görüyor (değer fırsatı veya sorun sinyali).\n"
        "Öneri: Bankalar ve sigorta şirketleri için en anlamlı değerleme metriğidir."
    ),
    (
        "F/S (P/S)", "priceToSalesTrailing12Months", None, '#,##0.00',
        "Fiyat / Satış oranı — son 12 aylık gelire göre.\n"
        "Öneri: Henüz kâr etmeyen büyüme şirketleri için F/K yerine F/S kullanın.\n"
        "Kural: Sektöre göre değişir; genel olarak F/S < 2 makul kabul edilir."
    ),
    (
        "FD/FAVÖK", "enterpriseToEbitda", None, '#,##0.00',
        "Firma Değeri / FAVÖK — sermaye yapısından bağımsız değerleme oranı.\n"
        "Yorum: Düşük FD/FAVÖK ucuzluğa; yüksek değer pahalılığa işaret eder (sektöre göre değişir).\n"
        "Öneri: Çok borçlu şirketleri karşılaştırırken F/K yerine bu oranı tercih edin."
    ),
    (
        "FD/Gelir", "enterpriseToRevenue", None, '#,##0.00',
        "Firma Değeri / Toplam Gelir oranı.\n"
        "Öneri: Kârsız veya düşük kârlı şirketleri değerlerken F/S yerine bu oran sermaye maliyetini de içerdiğinden daha kapsamlıdır."
    ),

    # ── 4. KARLILIK ───────────────────────────────────────────────────────────
    (
        "Brüt Kar Marjı %", "grossMargins", "yuzde", '#,##0.00"%"',
        "Brüt Kâr / Toplam Gelir — üretim/hizmet maliyeti düşüldükten sonra kalan pay.\n"
        "Yorum: Yüksek marj → güçlü fiyat belirleme gücü veya düşük maliyet yapısı.\n"
        "Öneri: Aynı sektördeki rakiplerle kıyaslayın; marjın zaman içinde daralıp daralmadığını izleyin."
    ),
    (
        "Faaliyet Marjı %", "operatingMargins", "yuzde", '#,##0.00"%"',
        "Faaliyet Kârı / Toplam Gelir — amortisman ve genel giderler düşüldükten sonraki kârlılık.\n"
        "Öneri: Brüt marjdan belirgin şekilde düşükse yönetim/genel giderler yüksek demektir; verimlilik sorusu işaret eder."
    ),
    (
        "Net Kar Marjı %", "profitMargins", "yuzde", '#,##0.00"%"',
        "Net Kâr / Toplam Gelir — vergi ve faiz sonrası nihai kârlılık.\n"
        "Öneri: Faaliyet marjıyla arasındaki fark büyükse finansman maliyetleri veya vergi yükü yüksektir.\n"
        "Kural: Tutarlı pozitif net marj güçlü iş modelinin göstergesidir."
    ),
    (
        "FAVÖK Marjı %", "ebitdaMargins", "yuzde", '#,##0.00"%"',
        "FAVÖK / Toplam Gelir — amortisman ve faiz öncesi nakit kârlılık.\n"
        "Öneri: Sermaye yoğun sektörlerde (enerji, altyapı) net marj yerine FAVÖK marjını tercih edin.\n"
        "Not: Amortisman politikasından etkilenmez; şirketler arası karşılaştırmayı kolaylaştırır."
    ),
    (
        "ROA %", "returnOnAssets", "yuzde", '#,##0.00"%"',
        "Varlık Kârlılığı: Net Kâr / Toplam Varlıklar × 100.\n"
        "Yorum: Şirketin sahip olduğu her 100 TL'lik varlıktan ne kadar kâr ürettiğini gösterir.\n"
        "Öneri: ROA > %5 genel olarak iyi kabul edilir; sektöre göre beklenti değişir.\n"
        "Kural: Bankalar için ROA < %1 bile normal olabilir; sektör ortalamasıyla karşılaştırın."
    ),
    (
        "ROE %", "returnOnEquity", "yuzde", '#,##0.00"%"',
        "Özkaynak Kârlılığı: Net Kâr / Özkaynak × 100.\n"
        "Yorum: Hissedarların yatırımından elde edilen getiriyi gösterir.\n"
        "Öneri: ROE > %15 güçlü; ancak yüksek borçlanma ROE'yi yapay olarak şişirebilir.\n"
        "Kural: ROE'yi Borç/Özkaynak oranıyla birlikte değerlendirin."
    ),

    # ── 5. BÜYÜME ─────────────────────────────────────────────────────────────
    (
        "Gelir Büyümesi %", "revenueGrowth", "yuzde", '#,##0.00"%"',
        "Son çeyreğin yıllık gelir büyümesi (YoY).\n"
        "Öneri: Enflasyon ortamında reel büyümeyi görmek için TÜFE'yi çıkarın.\n"
        "Kural: Sürekli çift haneli büyüme güçlü iş modeline işaret eder."
    ),
    (
        "Kazanç Büyümesi %", "earningsGrowth", "yuzde", '#,##0.00"%"',
        "Son çeyreğin yıllık kazanç (EPS) büyümesi (YoY).\n"
        "Öneri: Gelir büyümesinden yüksekse verimlilik artışı var demektir; düşükse maliyet baskısı sorgulanmalıdır."
    ),
    (
        "Çeyreklik Kazanç B. %", "earningsQuarterlyGrowth", "yuzde", '#,##0.00"%"',
        "Son çeyreğin bir önceki yılın aynı çeyreğine göre kazanç büyümesi.\n"
        "Öneri: Mevsimsellik etkisini elemek için QoQ yerine YoY karşılaştırma olan bu metriği tercih edin.\n"
        "Not: Tek bir çeyreğe odaklanmak yanıltıcı olabilir; 4–8 çeyreklik trende bakın."
    ),

    # ── 6. GELİR TABLOSU ──────────────────────────────────────────────────────
    (
        "Toplam Gelir", "totalRevenue", None, '#,##0',
        "Son 12 aylık (TTM) toplam satış geliri (TL).\n"
        "Öneri: Tek başına yeterli değildir; kâr marjlarıyla birlikte değerlendirin.\n"
        "Not: Holding şirketlerinde iştirak gelirleri dahil olmayabilir."
    ),
    (
        "Brüt Kar", "grossProfits", None, '#,##0',
        "Toplam Gelir − Satılan Malın Maliyeti (TL).\n"
        "Öneri: Brüt kâr trendi, fiyatlama gücünün ve maliyet kontrolünün en net göstergesidir."
    ),
    (
        "FAVÖK", "ebitda", None, '#,##0',
        "Faiz, Amortisman ve Vergi Öncesi Kâr (TL).\n"
        "Öneri: Borç ödeme kapasitesini değerlendirmek için Toplam Borç / FAVÖK oranını hesaplayın.\n"
        "Kural: Borç / FAVÖK > 4 aşırı kaldıraç; < 2 sağlıklı bilanço olarak kabul edilir."
    ),
    (
        "Net Gelir", "netIncomeToCommon", None, '#,##0',
        "Adi hissedarlara düşen net kâr (TL) — imtiyazlı temettü sonrası.\n"
        "Öneri: Nakit akışıyla birlikte değerlendirin; yüksek net kâra rağmen negatif nakit akışı yaratıcı muhasebe sorusunu doğurur."
    ),
    (
        "HBK — Trailing EPS", "trailingEps", None, '#,##0.00',
        "Hisse Başına Kazanç — son 12 aylık gerçekleşen (TL/hisse).\n"
        "Öneri: F/K hesaplamak için Güncel Fiyat ÷ Trailing EPS formülünü kullanın.\n"
        "Not: Hisse geri alımları EPS'i artırır; gerçek kâr büyümesiyle karıştırmayın."
    ),
    (
        "HBK — Forward EPS", "forwardEps", None, '#,##0.00',
        "Hisse Başına Kazanç — önümüzdeki 12 aylık analist tahmini (TL/hisse).\n"
        "Öneri: Trailing EPS ile karşılaştırarak analistlerin büyüme beklentisini görün.\n"
        "Not: Tahmine dayalıdır; şirkete özel riskler gerçekleşirse sapma olabilir."
    ),

    # ── 7. BİLANÇO & FİNANSAL SAĞLIK ────────────────────────────────────────
    (
        "Toplam Nakit", "totalCash", None, '#,##0',
        "Nakit ve nakit benzeri varlıkların toplamı (TL).\n"
        "Öneri: Toplam Nakit / Piyasa Değeri oranı yüksekse şirket ucuz veya savunmacı bir bilançoya sahip olabilir."
    ),
    (
        "Hisse Başı Nakit", "totalCashPerShare", None, '#,##0.00',
        "Hisse başına düşen nakit miktarı (TL/hisse).\n"
        "Öneri: Güncel fiyatın önemli bir kısmı nakitse düşük değerleme veya potansiyel temettü artışı sinyali olabilir."
    ),
    (
        "Toplam Borç", "totalDebt", None, '#,##0',
        "Kısa ve uzun vadeli finansal borçların toplamı (TL).\n"
        "Öneri: Toplam Borç / FAVÖK ve Borç / Özkaynak oranlarıyla birlikte okuyun.\n"
        "Kural: Yüksek faiz ortamında borçlu şirketler daha fazla baskı altında kalır."
    ),
    (
        "Borç / Özkaynak", "debtToEquity", None, '#,##0.00',
        "Toplam Borç / Özkaynak — kaldıraç oranı.\n"
        "Yorum: Yüksek oran finansal risk anlamına gelir; sektöre göre normal eşik değişir.\n"
        "Öneri: Bankalar ve finans şirketleri için bu oran doğası gereği yüksektir; sektör karşılaştırması yapın.\n"
        "Kural: İmalat sektörü için >2 riskli; <1 muhafazakâr kabul edilir."
    ),
    (
        "Cari Oran", "currentRatio", None, '#,##0.00',
        "Dönen Varlıklar / Kısa Vadeli Yükümlülükler — kısa vadeli likidite göstergesi.\n"
        "Yorum: >1 → kısa vadeli borçları karşılayabilir; <1 → likidite riski var.\n"
        "Öneri: 1.5–2.5 arası sağlıklı kabul edilir; çok yüksek cari oran atıl varlık sorusunu doğurabilir."
    ),
    (
        "Asit-Test Oranı", "quickRatio", None, '#,##0.00',
        "(Dönen Varlıklar − Stoklar) / Kısa Vadeli Yükümlülükler — daha katı likidite testi.\n"
        "Öneri: Cari orandan düşükse stok bağımlılığı yüksektir; perakende ve imalat için özellikle önemlidir.\n"
        "Kural: >1 güvenli; <0.5 acil likidite sorunu sinyali verebilir."
    ),
    (
        "Defter Değeri / Hisse", "bookValue", None, '#,##0.00',
        "Hisse başına özkaynak defter değeri (TL/hisse).\n"
        "Öneri: Güncel Fiyat / Defter Değeri = F/DD oranını kendiniz hesaplayarak kontrol edin.\n"
        "Not: Enflasyon muhasebesi uygulayan şirketlerde defter değeri önemli ölçüde değişebilir."
    ),

    # ── 8. NAKİT AKIŞI ───────────────────────────────────────────────────────
    (
        "Serbest Nakit Akışı", "freeCashflow", None, '#,##0',
        "Faaliyet Nakit Akışı − Sermaye Harcamaları (TL).\n"
        "Öneri: Pozitif ve büyüyen serbest nakit akışı; kâr kalitesinin en güvenilir göstergesidir.\n"
        "Kural: Net kâr pozitif ama serbest nakit akışı negatifse muhasebe kalitesini sorgulayın."
    ),
    (
        "Faaliyet Nakit Akışı", "operatingCashflow", None, '#,##0',
        "Ana faaliyetlerden elde edilen net nakit girişi (TL).\n"
        "Öneri: Net kâr ile faaliyet nakit akışı arasındaki sürekli büyük fark, tahakkuk temelli gelir tanıma riskine işaret edebilir."
    ),

    # ── 9. TEMETTÜ ────────────────────────────────────────────────────────────
    (
        "Temettü (Yıllık)", "dividendRate", None, '#,##0.00',
        "Hisse başına yıllık brüt temettü tutarı (TL/hisse).\n"
        "Öneri: Güncel fiyata bölerek temettü verimini kendiniz doğrulayın.\n"
        "Not: Türk şirketleri genellikle yılda bir temettü öder."
    ),
    (
        "Temettü Verimi %", "dividendYield", "yuzde", '#,##0.00"%"',
        "Yıllık Temettü / Güncel Fiyat × 100.\n"
        "Öneri: Mevduat faiz oranlarıyla karşılaştırın; temettü verimi faizin belirgin altındaysa risk primini sorgulayın.\n"
        "Kural: Sürdürülebilir yüksek temettü için Ödeme Oranı'nın %80'in altında olması tercih edilir."
    ),
    (
        "Ödeme Oranı %", "payoutRatio", "yuzde", '#,##0.00"%"',
        "Temettü / Net Kâr × 100 — kârın ne kadarının dağıtıldığını gösterir.\n"
        "Yorum: %100'ün üzeri → kârdan fazla temettü ödenmiş; sürdürülemez.\n"
        "Öneri: %40–70 arası büyüme ve temettüyü dengeleyen sağlıklı bir aralık olarak değerlendirilebilir."
    ),
    (
        "5Y Ort. Temettü %", "fiveYearAvgDividendYield", None, '#,##0.00"%"',
        "Son 5 yılın ortalama temettü verimi.\n"
        "Öneri: Güncel temettü verimi 5 yıllık ortalamanın belirgin üzerindeyse hisse fiyatı düşmüş veya temettü artmış olabilir; sebebi araştırın."
    ),
    (
        "Son Temettü Değeri", "lastDividendValue", None, '#,##0.00',
        "En son ödenen temettü tutarı (TL/hisse).\n"
        "Öneri: Önceki dönemle karşılaştırarak temettü artış trendini değerlendirin."
    ),
    (
        "Ex-Temettü Tarihi", "exDividendDate", "tarih", None,
        "Bu tarihten önce alınan hisseler temettüye hak kazanır.\n"
        "Öneri: Ex-temettü gününde fiyat genellikle temettü tutarı kadar düşer; bunu fırsat veya tuzak olarak değerlendirirken dikkatli olun."
    ),

    # ── 10. ANALİST & DİĞER ──────────────────────────────────────────────────
    (
        "Analist Tavsiyesi", "recommendationKey", None, None,
        "Analistlerin konsensüs tavsiyesi (strong_buy / buy / hold / sell / strong_sell).\n"
        "Öneri: Tek başına güvenilir değildir; analist sayısı azsa daha az anlamlıdır.\n"
        "Not: Tavsiyeler geriye dönük olarak güncellenebilir."
    ),
    (
        "Hedef Fiyat (Ort.)", "targetMeanPrice", None, '#,##0.00',
        "Analistlerin 12 aylık ortalama hedef fiyat tahmini (TL).\n"
        "Öneri: (Hedef Fiyat − Güncel Fiyat) / Güncel Fiyat = Beklenen Getiri potansiyelini hesaplayın.\n"
        "Not: Analist sayısı azsa (< 3) bu değere temkinli yaklaşın."
    ),
    (
        "Hedef Fiyat (Yük.)", "targetHighPrice", None, '#,##0.00',
        "Analistlerin en iyimser 12 aylık hedef fiyat tahmini (TL).\n"
        "Öneri: Yüksek ve düşük hedef arasındaki geniş fark, analistler arasında yüksek belirsizliğe işaret eder."
    ),
    (
        "Hedef Fiyat (Düş.)", "targetLowPrice", None, '#,##0.00',
        "Analistlerin en kötümser 12 aylık hedef fiyat tahmini (TL).\n"
        "Öneri: Güncel fiyat en düşük hedefin de altındaysa analistlerin tamamına göre aşırı satış durumu olabilir."
    ),
    (
        "Analist Sayısı", "numberOfAnalystOpinions", None, '#,##0',
        "Hedef fiyat ve tavsiye oluşturan analist sayısı.\n"
        "Öneri: < 3 analist → konsensüs güvenilirliği düşük; > 10 analist → daha sağlıklı konsensüs."
    ),
    (
        "Short Oranı", "shortRatio", None, '#,##0.00',
        "Açık pozisyon adedinin günlük ortalama hacme bölümü — açık pozisyonların kapanması için gereken gün sayısı.\n"
        "Yorum: Yüksek short oranı → piyasanın düşüş beklentisi; ancak sıkışma (short squeeze) potansiyeli de taşır.\n"
        "Öneri: > 10 gün ise dikkatli olun; short squeeze riski hem fırsat hem tehlike yaratabilir."
    ),
    (
        "Short % Float", "shortPercentOfFloat", "yuzde", '#,##0.00"%"',
        "Açık pozisyondaki hisse adedinin dolaşımdaki hisseye oranı.\n"
        "Öneri: > %20 yüksek açık pozisyon; short squeeze beklentisiyle spekülatif alım artabilir.\n"
        "Kural: > %30 aşırı açık; fiyat hareketi çok sert olabilir."
    ),
    (
        "İçeriden Pay %", "heldPercentInsiders", "yuzde", '#,##0.00"%"',
        "Yönetici ve büyük hissedarların elindeki pay oranı.\n"
        "Yorum: Yüksek içeriden sahiplik yönetimin şirketle çıkar birliğini gösterir (olumlu sinyal).\n"
        "Öneri: İçeriden alım/satım işlemlerini de takip edin; sahiplik oranındaki değişimler önemlidir."
    ),
    (
        "Kurumsal Pay %", "heldPercentInstitutions", "yuzde", '#,##0.00"%"',
        "Fonlar, emeklilik şirketleri vb. kurumsal yatırımcıların elindeki pay oranı.\n"
        "Yorum: Yüksek kurumsal sahiplik genellikle profesyonel analiz süzgecinden geçtiğine işaret eder.\n"
        "Öneri: > %70 ise kurumsal çıkış fiyat üzerinde baskı yaratabilir; bant genişliğini izleyin."
    ),
]

# Kategori adı, kapsadığı başlıklar, renk
KATEGORILER = [
    ("Kimlik & Şirket",
     ["Hisse Kodu","Şirket İsmi","Sektör","Sektör Altı","Ülke","Çalışan Sayısı","Borsa"],
     "2C3E50"),
    ("Fiyat & Piyasa",
     ["Güncel Fiyat","Piyasa Değeri","Firma Değeri","Dolaşımdaki Hisse","Toplam Hisse Sayısı",
      "52H Yüksek","52H Düşük","50G Ortalama","200G Ortalama","Ort. Hacim (3A)","Ort. Hacim (10G)","Beta"],
     "1A5276"),
    ("Değerleme Oranları",
     ["F/K (Trailing)","F/K (Forward)","PEG Oranı","F/DD (P/B)","F/S (P/S)","FD/FAVÖK","FD/Gelir"],
     "154360"),
    ("Karlılık",
     ["Brüt Kar Marjı %","Faaliyet Marjı %","Net Kar Marjı %","FAVÖK Marjı %","ROA %","ROE %"],
     "1E5631"),
    ("Büyüme",
     ["Gelir Büyümesi %","Kazanç Büyümesi %","Çeyreklik Kazanç B. %"],
     "4A235B"),
    ("Gelir Tablosu",
     ["Toplam Gelir","Brüt Kar","FAVÖK","Net Gelir","HBK — Trailing EPS","HBK — Forward EPS"],
     "784212"),
    ("Bilanço & Fin. Sağlık",
     ["Toplam Nakit","Hisse Başı Nakit","Toplam Borç","Borç / Özkaynak",
      "Cari Oran","Asit-Test Oranı","Defter Değeri / Hisse"],
     "1A5276"),
    ("Nakit Akışı",
     ["Serbest Nakit Akışı","Faaliyet Nakit Akışı"],
     "117A65"),
    ("Temettü",
     ["Temettü (Yıllık)","Temettü Verimi %","Ödeme Oranı %","5Y Ort. Temettü %",
      "Son Temettü Değeri","Ex-Temettü Tarihi"],
     "7B241C"),
    ("Analist & Diğer",
     ["Analist Tavsiyesi","Hedef Fiyat (Ort.)","Hedef Fiyat (Yük.)","Hedef Fiyat (Düş.)",
      "Analist Sayısı","Short Oranı","Short % Float","İçeriden Pay %","Kurumsal Pay %"],
     "212F3D"),
]
# ──────────────────────────────────────────────────────────────────────────────


def hisse_listesini_getir(dosya_adi: str) -> list:
    """Lokal bir metin dosyasından hisse kodlarını okur (her satırda bir kod)."""
    print(f"📁 Hisse listesi '{dosya_adi}' dosyasından okunuyor…")
    try:
        with open(dosya_adi, "r", encoding="utf-8") as f:
            kodlar = [satir.strip() for satir in f if satir.strip()]
        print(f"✅ {len(kodlar)} hisse kodu bulundu.\n")
        return kodlar
    except FileNotFoundError:
        print(f"❌ Dosya bulunamadı: {dosya_adi}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Dosya okunamadı: {e}")
        sys.exit(1)


def yuzde_donusum(deger):
    if deger is None:
        return None
    try:
        return round(float(deger) * 100, 4)
    except Exception:
        return None


def tarih_donusum(deger):
    if deger is None:
        return None
    try:
        return datetime.fromtimestamp(int(deger), tz=timezone.utc).strftime("%d.%m.%Y")
    except Exception:
        return str(deger)


def temizle_kod(kod: str) -> str:
    """GARAN.IS → GARAN"""
    return kod.replace(".IS", "").replace(".is", "")


def hisse_verisi_getir(kod_is: str) -> dict:
    """
    kod_is  : yfinance için .IS uzantılı kod (örn. GARAN.IS)
    Hisse Kodu sütununa .IS olmadan yazar.
    """
    satir = {}
    try:
        ticker = yf.Ticker(kod_is)
        info   = ticker.info

        if not info.get("currentPrice"):
            info["currentPrice"] = info.get("regularMarketPrice")

        for baslik, anahtar, on_islem, _, _ in SUTUNLAR:
            if anahtar is None:                         # "ozel" → Hisse Kodu
                satir[baslik] = temizle_kod(kod_is)
                continue
            ham = info.get(anahtar)
            if on_islem == "yuzde":
                satir[baslik] = yuzde_donusum(ham)
            elif on_islem == "tarih":
                satir[baslik] = tarih_donusum(ham)
            else:
                satir[baslik] = ham

    except Exception as e:
        print(f"  ⚠️  {kod_is}: {e}")
        for baslik, _, _, _, _ in SUTUNLAR:
            satir.setdefault(baslik, None)
        satir["Hisse Kodu"]  = temizle_kod(kod_is)
        satir["Şirket İsmi"] = "HATA"

    return satir


def verileri_topla(kodlar: list) -> pd.DataFrame:
    satirlar = []
    toplam   = len(kodlar)
    for i, kod in enumerate(kodlar, 1):
        print(f"[{i:>{len(str(toplam))}}/{toplam}]  {temizle_kod(kod)} işleniyor…",
              end=" ", flush=True)
        veri = hisse_verisi_getir(kod)
        satirlar.append(veri)
        print("✗" if veri.get("Şirket İsmi") == "HATA" else "✓")
        time.sleep(ISTEK_ARASI_BEKLEME)
    sutun_sirasi = [s[0] for s in SUTUNLAR]
    return pd.DataFrame(satirlar, columns=sutun_sirasi)


# ─── EXCEL FORMATLAMA ─────────────────────────────────────────────────────────

ACIK_MAVI = "D6E4F0"
BEYAZ     = "FFFFFF"
HATA_BG   = "FFE0E0"
NEG_RENK  = "C00000"
NOT_YAZAR = "Temel Veri Çekici"


def _kenar(renk="BFBFBF"):
    t = Side(style="thin", color=renk)
    return Border(left=t, right=t, top=t, bottom=t)


def _kategori_hucre(ws, satir, bas, bit, metin, renk):
    ws.merge_cells(start_row=satir, start_column=bas, end_row=satir, end_column=bit)
    h = ws.cell(satir, bas)
    h.value     = metin
    h.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    h.fill      = PatternFill("solid", start_color=renk)
    h.alignment = Alignment(horizontal="center", vertical="center")


def _alan_basligi(hucre, metin, renk, not_metni: str):
    hucre.value     = metin
    hucre.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    hucre.fill      = PatternFill("solid", start_color=renk)
    hucre.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hucre.border    = _kenar("7F8C8D")
    if not_metni:
        yorum = Comment(not_metni, NOT_YAZAR)
        yorum.width  = 360
        yorum.height = 160
        hucre.comment = yorum


def _veri(hucre, cift: bool, hata: bool = False):
    renk = HATA_BG if hata else (ACIK_MAVI if cift else BEYAZ)
    hucre.fill      = PatternFill("solid", start_color=renk)
    hucre.font      = Font(name="Arial", size=9)
    hucre.alignment = Alignment(horizontal="center", vertical="center")
    hucre.border    = _kenar()


def excel_kaydet(df: pd.DataFrame, dosya: str):
    print(f"\n💾  Excel oluşturuluyor: {dosya}")

    df.to_excel(dosya, index=False, sheet_name="Temel Veriler", startrow=2)

    wb = load_workbook(dosya)
    ws = wb["Temel Veriler"]

    basliklar = [s[0] for s in SUTUNLAR]
    formatlar  = {s[0]: s[3] for s in SUTUNLAR}
    notlar     = {s[0]: s[4] for s in SUTUNLAR}

    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 0
    ws.row_dimensions[3].height = 36

    baslik_renk = {}
    for _, kat_basliklar, renk in KATEGORILER:
        for b in kat_basliklar:
            baslik_renk[b] = renk

    for sutun_no, baslik in enumerate(basliklar, 1):
        harf = get_column_letter(sutun_no)
        ws.column_dimensions[harf].width = max(12, min(len(baslik) + 3, 24))
        hucre = ws.cell(3, sutun_no)
        renk  = baslik_renk.get(baslik, "2C3E50")
        _alan_basligi(hucre, baslik, renk, notlar.get(baslik, ""))

    sutun_idx = {b: i + 1 for i, b in enumerate(basliklar)}
    for kat_ad, kat_basliklar, renk in KATEGORILER:
        mevcut = [b for b in kat_basliklar if b in sutun_idx]
        if not mevcut:
            continue
        bas = sutun_idx[mevcut[0]]
        bit = sutun_idx[mevcut[-1]]
        _kategori_hucre(ws, 1, bas, bit, kat_ad, renk)

    for sutun_no in range(1, len(basliklar) + 1):
        h = ws.cell(2, sutun_no)
        h.value = None
        h.fill  = PatternFill("solid", start_color="EEEEEE")

    for satir_no in range(4, ws.max_row + 1):
        cift = satir_no % 2 == 0
        hata = ws.cell(satir_no, 2).value == "HATA"

        for sutun_no, baslik in enumerate(basliklar, 1):
            hucre = ws.cell(satir_no, sutun_no)
            _veri(hucre, cift, hata)

            fmt = formatlar.get(baslik)
            if fmt:
                hucre.number_format = fmt

            if "%" in baslik and isinstance(hucre.value, (int, float)) and hucre.value < 0:
                hucre.font = Font(name="Arial", size=9, color=NEG_RENK)

    ws.freeze_panes = "B4"
    ws.sheet_properties.tabColor = "1A5276"

    ws_b = wb.create_sheet("Bilgi")
    bilgiler = [
        ("Oluşturulma Tarihi",   datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("Veri Kaynağı",         "Yahoo Finance (yfinance)"),
        ("Toplam Sütun",         len(basliklar)),
        ("Toplam Hisse",         len(df)),
        ("Başarılı",             len(df[df["Şirket İsmi"] != "HATA"])),
        ("Başarısız / Veri Yok", len(df[df["Şirket İsmi"] == "HATA"])),
    ]
    for i, (a, b) in enumerate(bilgiler, 1):
        ws_b.cell(i, 1, a).font = Font(bold=True, name="Arial")
        ws_b.cell(i, 2, b).font = Font(name="Arial")
    ws_b.column_dimensions["A"].width = 25
    ws_b.column_dimensions["B"].width = 35

    wb.save(dosya)
    print(f"✅  Kaydedildi → {dosya}")


# ─── ANA AKIŞ ─────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  📊  Hisse Senedi Temel Veri Çekici — GitHub Actions Sürümü")
    print("=" * 62)

    # 1. Hisse kodlarını dosyadan oku
    kodlar = hisse_listesini_getir(HISSE_LISTESI_DOSYASI)

    # 2. yfinance için .IS uzantısını ekle (dosyada yoksa)
    kodlar_is = [k if k.endswith(".IS") else f"{k}.IS" for k in kodlar]

    # 3. Verileri topla
    df = verileri_topla(kodlar_is)

    # 4. Dinamik Excel dosya adı (Türkiye saati ile gün_ay_yıl)
    turkey_tz = timezone(timedelta(hours=3))
    bugun = datetime.now(turkey_tz)
    tarih_str = bugun.strftime("%d_%m_%Y")   # örnek: 21_03_2025
    
    gunluk_dosya = f"hisse_temel_veriler_{tarih_str}.xlsx"
    sabit_dosya = f"hisse_temel_veriler.xlsx"

    excel_kaydet(df, gunluk_dosya)
    excel_kaydet(df, sabit_dosya)
    
    # 6. Özet bilgi
    basarili  = len(df[df["Şirket İsmi"] != "HATA"])
    basarisiz = len(df) - basarili
    basliklar = [s[0] for s in SUTUNLAR]

    print("\n─── ÖZET ────────────────────────────────────────────────")
    print(f"  Toplam hisse   : {len(df)}")
    print(f"  Başarılı       : {basarili}")
    print(f"  Veri alınamadı : {basarisiz}")
    print(f"  Toplam sütun   : {len(basliklar)}")
    print(f"  Çıktı dosyası  : {gunluk_dosya}")
    print(f"  Çıktı dosyası  : {sabit_dosya}")
    print("─────────────────────────────────────────────────────────\n")


if __name__ == "__main__":
    main()